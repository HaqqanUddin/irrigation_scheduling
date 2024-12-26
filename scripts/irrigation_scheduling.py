# Import necessary libraries
import pandas as pd  # For data manipulation and reading from Excel
from openpyxl import Workbook  # For writing data to Excel with formatting
from openpyxl.styles import Font  # For styling Excel headers
import datetime  # For date handling in interpolation
import numpy as np  # For numerical operations, especially generating daily values from monthly data

# Function to calculate crop evapotranspiration (ET) based on crop coefficient (kc) and reference ET (ETr)
def calculate_et(crop_kc, etr):
    """
    Calculates the crop evapotranspiration (ETc) using the formula: ETc = ETr * kc
    
    Parameters:
    - crop_kc: Crop coefficient for the current growth stage
    - etr: Reference evapotranspiration (ETr) in mm/day
    
    Returns:
    - ETc: Crop evapotranspiration (ETc) in mm/day
    """
    return etr * crop_kc

# Function to interpolate monthly data to daily values using linear interpolation
def interpolate_monthly_to_daily(monthly_data, season_months):
    """
    Interpolates monthly data to daily values for the growing season using linear interpolation.
    
    Parameters:
    - monthly_data: List of monthly data (e.g., ETr or rainfall)
    - season_months: List of integers representing months in the growing season
    
    Returns:
    - daily_data: List of interpolated daily data values for the entire growing season
    """
    daily_data = []
    for i in range(len(season_months) - 1):
        start = monthly_data[i]
        end = monthly_data[i + 1]
        # Calculate the number of days in the month
        days_in_month = (datetime.date(2024, season_months[i + 1], 1) - datetime.date(2024, season_months[i], 1)).days
        # Generate daily values within the month using linear interpolation
        daily_values = np.linspace(start, end, days_in_month, endpoint=False)
        daily_data.extend(daily_values)
    # Handle the transition from the last month to the next year (December to the first month)
    start = monthly_data[-1]
    end = monthly_data[0]
    days_in_month = (datetime.date(2024, 12, 31) - datetime.date(2024, season_months[-1], 1)).days + 1
    daily_values = np.linspace(start, end, days_in_month, endpoint=False)
    daily_data.extend(daily_values[:days_in_month])  # Ensure this slice does not exceed the intended range
    return daily_data

# Main function to generate daily irrigation schedule for the growing season
def daily_irrigation_schedule(soil, crop, climate, season_months, rainfall_data):
    """
    Generates a daily irrigation schedule for the crop based on soil, crop, and climate data.
    
    Parameters:
    - soil: Dictionary containing soil properties (e.g., field capacity, wilting point)
    - crop: Dictionary containing crop properties (e.g., growth stages, crop coefficient)
    - climate: Dictionary containing monthly ETr and rainfall data
    - season_months: List of months in the growing season
    - rainfall_data: List of interpolated daily rainfall data
    
    Returns:
    - schedule: A list of daily irrigation schedule records
    """
    schedule = []
    # Interpolate ETr and rainfall data from monthly to daily values
    daily_etr = interpolate_monthly_to_daily(climate['ETr'], season_months)
    daily_rainfall = interpolate_monthly_to_daily(rainfall_data, season_months)
    
    day_count = 1  # Start the day count for the season
    cumulative_soil_water_deficit = 0  # Initialize cumulative soil water deficit
    max_allowable_depletion = 0.7 * (soil['field_capacity'] - soil['wilting_point']) * crop['root_depth'] * 10  # Max allowable soil depletion

    # Loop through each crop growth stage (e.g., initial, development, mid-season, late-season)
    for stage, properties in crop['growth_stages'].items():
        days = properties['days']  # Days in the current growth stage
        kc = properties['kc']  # Crop coefficient for the current stage
        
        # Process each day in the current growth stage
        for day in range(days):
            daily_climate = {
                'ETr': daily_etr[day_count - 1],  # Get the daily ETr value
                'Rainfall': daily_rainfall[day_count - 1]  # Get the daily rainfall value
            }

            # Calculate crop water use (ETc)
            etc = calculate_et(kc, daily_climate['ETr'])

            # Calculate effective rainfall (assuming no losses)
            effective_rainfall = max(daily_climate['Rainfall'] - 0, 0)  # No losses in this example

            # Determine net irrigation requirement
            soil_water_deficit = effective_rainfall - etc

            # Update the cumulative soil water deficit
            cumulative_soil_water_deficit += soil_water_deficit

            # Check if irrigation is needed
            if cumulative_soil_water_deficit > max_allowable_depletion:
                irrigation_required = True
                daily_irrigation_req = max_allowable_depletion  # Apply maximum allowable irrigation
                applied_water = daily_irrigation_req
                cumulative_soil_water_deficit = 0  # Reset soil water deficit after irrigation
            else:
                irrigation_required = False
                daily_irrigation_req = 0  # No irrigation needed
                applied_water = 0

            # Append the daily irrigation data to the schedule
            schedule.append({
                'day': day_count,  # Day of the season
                'growing_season': 'Kharif',  # Kharif season in this example
                'growth_stage': stage,  # Current growth stage
                'crop_type': crop['crop_type'],  # Crop type (e.g., cotton)
                'kc': kc,  # Crop coefficient
                'ETo (mm/day)': daily_climate['ETr'],  # Reference evapotranspiration (ETr)
                'Crop water use (Etc) (mm/day)': etc,  # Crop evapotranspiration (ETc)
                'Rainfall (mm)': daily_climate['Rainfall'],  # Daily rainfall
                'Net Irrigation application (mm)': round(daily_irrigation_req, 2),  # Net irrigation required
                'Cumulative soil water deficit (mm)': round(cumulative_soil_water_deficit, 2),  # Soil water deficit
                'Irrigation Required': 'Yes' if irrigation_required else 'No'  # Whether irrigation is required
            })
            day_count += 1  # Increment the day count
    return schedule  # Return the generated irrigation schedule

# Correct file path for reading Excel data
file_path = r"D:\E\Master Courses\Semesters\Fifth Semester\CE-577 Irrigation System Design and Management\Term Project 2\Climatic_Data.xlsx"

# Read soil, crop, and climatic data from the Excel file
soil_data = pd.read_excel(file_path, sheet_name="Soil")
crop_data = pd.read_excel(file_path, sheet_name="Crops")
climatic_data = pd.read_excel(file_path, sheet_name="Climate")

# Handle potential missing sheet or columns for rainfall data
try:
    rainfall_data = pd.read_excel(file_path, sheet_name="Rainfall")
except KeyError:
    print("Error: 'Rainfall' sheet or column not found in Excel file.")
    exit()

# Ask user for input to specify soil type and crop type
soil_type = input("Enter the soil type (sand, sandy loam, loam, clay loam, silty clay, clay): ").strip().lower()
crop_type = input("Enter the crop type (cotton, sugarcane, rice, maize, wheat): ").strip().lower()

# Extract soil and crop properties based on user input
soil = soil_data[soil_data['soil_type'] == soil_type].iloc[0].to_dict()
crop = crop_data[crop_data['crop_type'] == crop_type].iloc[0].to_dict()

# Define crop growth stages with crop coefficient (kc) and duration
crop['growth_stages'] = {
    'initial': {'kc': crop['initial_kc'], 'days': crop['initial_days']},
    'development': {'kc': crop['development_kc'], 'days': crop['development_days']},
    'mid-season': {'kc': crop['mid_season_kc'], 'days': crop['mid_season_days']},
    'late-season': {'kc': crop['late_season_kc'], 'days': crop['late_season_days']}
}

# Define the months for the Kharif season (April to September)
season_months = [4, 5, 6, 7, 8, 9]

# Extract climatic data for the Kharif season
monthly_etr = climatic_data['ETr'][0:6].tolist()
monthly_rainfall = rainfall_data['Rainfall (mm)'][0:6].tolist()

# Combine ETr and rainfall data into a single climatic conditions dictionary
climatic_conditions = {
    'ETr': monthly_etr,
    'Rainfall': monthly_rainfall
}

# Generate the irrigation schedule
schedule = daily_irrigation_schedule(soil, crop, climatic_conditions, season_months, monthly_rainfall)

# Create a DataFrame from the schedule for easier handling
schedule_df = pd.DataFrame(schedule)

# Correct file path for saving the output schedule
output_file = r"D:\E\Master Courses\Semesters\Fifth Semester\CE-577 Irrigation System Design and Management\Term Project 2\Irrigation_Schedule.xlsx"

# Set up the workbook and worksheet for output
wb = Workbook()
ws = wb.active

# Write the headers to the Excel sheet
headers = ['Day', 'Growing Season', 'Growth Stage', 'Crop Type', 'Kc', 'ETo (mm/day)', 'Crop water use (Etc) (mm/day)', 'Rainfall (mm)', 'Net Irrigation application (mm)', 'Cumulative soil water deficit (mm)', 'Irrigation Required']
for col_idx, header in enumerate(headers):
    ws.cell(row=1, column=col_idx+1, value=header)  # Write header values
    ws.cell(row=1, column=col_idx+1).font = Font(bold=True)  # Set header font to bold

# Write the irrigation schedule data starting from row 2
for idx, row in schedule_df.iterrows():
    ws.append([
        row['day'],
        row['growing_season'],
        row['growth_stage'],
        row['crop_type'],
        row['kc'],
        round(row['ETo (mm/day)'], 2),  # Round values to 2 decimal places
        round(row['Crop water use (Etc) (mm/day)'], 2),
        round(row['Rainfall (mm)'], 2),
        round(row['Net Irrigation application (mm)'], 2),
        round(row['Cumulative soil water deficit (mm)'], 2),
        row['Irrigation Required']
    ])

# Save the workbook to the output file
wb.save(output_file)
print(f"Irrigation Schedule results saved to {output_file}")